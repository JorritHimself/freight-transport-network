#!C:\Python27
# -*- coding: utf-8 -*-
import sys
import time
from openpyxl import load_workbook
from modules import dijkstra, Graph

"""
    This module is meant to be visible for the user and to be used directly
    with no interaction with other subpackages of freight_transport_network.

    Provides a single class (Network) and main methods that use it to find
    shortest paths between all nodes of a railway network and a roadway
    network, represented by a list of links and its distances (see "data"
    folder) for each gauge used in the network.

    Inputs ("data" folder):
        railway_links_table.xlsx: List of railway links and their distances,
            for each different gauge.
        roadway_links_table.xlsx: List of railway links and their distances,
            for a unique gauge (roadway mode doesn't usually have "gauges").

    Outputs ("paths" folder):
        railway_shortest_paths.xlsx: List of paths between all nodes of the
            network, by gauge.
        roadway_shortest_paths.xlsx: List of paths between all nodes of the
            network, for a unique gauge.
"""


class Network():

    """Represents a network with different gauges.

    A Network instance can have many different gauges, wich in practice is
    like having different isolated subnetworks. Each gauge of the network has
    its own Graph representing all the nodes of the network-gauge connected by
    its links."""

    PATH_SHEET_SUFFIX = "paths"
    PATH_FIELDS = ["id_od", "origin", "destination", "distance", "path",
                   "gauge"]

    def __init__(self, gauge_names):
        self.gauge_names = gauge_names
        self.graphs = {}
        self.paths = {}

    # PUBLIC
    def create_graphs_from_excel(self, wb):
        """Create graphs from lists of links in excel."""

        # iterate through each gauge and worksheet
        for gauge in self.gauge_names:

            # take worksheet
            ws = wb.get_sheet_by_name(gauge)

            # create graph
            self.graphs[gauge] = Graph(ws)

    def calculate_paths(self):
        """Find shortest paths for each possible pair of nodes, by gauge."""

        # start total networks timer
        total_timer_start = time.time()

        # iterate through each gauge
        for gauge in self.gauge_names:

            # start network timer
            network_timer_start = time.time()

            # calculate paths for the network
            self._calculate_paths(gauge)

            # stop network timer
            elapsed = (time.time() - network_timer_start)
            self._report_time(elapsed, gauge)

        # stop total networks timer
        elapsed = (time.time() - total_timer_start)
        self._report_time(elapsed, "all networks")

    def store_paths_in_excel(self, wb, xl_output=None):
        """Store found paths in excel."""

        print "\n Saving results in excel..."
        sys.stdout.flush()

        # create worksheet to store all results
        ws_all = wb.create_sheet()
        ws_all.title = "all_" + self.PATH_SHEET_SUFFIX

        # write field names
        ws_all.append(self.PATH_FIELDS)

        # store paths for each network
        for gauge in self.paths:
            self._store_network_paths(wb, gauge, ws_all)

        # sort by distance before save
        # self._sort_by_distance(ws_all)

        # save workbook
        wb.save(xl_output or self.XL_OUTPUT)

        print "Finished."

    # PRIVATE
    def _calculate_paths(self, gauge):
        """Find shortest paths for each possible pair of nodes."""

        # take graph of network
        graph = self.graphs[gauge]

        # create paths dictionary
        self.paths[gauge] = {}
        paths = self.paths[gauge]

        # take nodes
        nodes = graph.keys()
        nodes.sort()

        # calcualte total paths
        total_paths = len(nodes) ** 2
        print total_paths, "paths will be calculated"

        # calculate minimum paths between all nodes
        index_calculated_paths = 0
        start_timer = time.time()

        for node_a in nodes:

            # create dictionary for node a in paths
            paths[node_a] = {}

            for node_b in nodes:

                # create dictionary for node b in node a
                paths[node_a][node_b] = {}

                # if is the same node, there is no path
                if node_a == node_b:

                    paths[node_a][node_b]["distance"] = 0.0
                    paths[node_a][node_b]["path"] = None

                else:

                    # use dijkstra to calculate minimum path from a to b
                    distance, path = dijkstra(graph, node_a, node_b)

                    # replace "a" in path for node_a
                    path[0] = node_a

                    # store results in paths
                    paths[node_a][node_b]["distance"] = distance
                    paths[node_a][node_b]["path"] = path

                # show progress
                index_calculated_paths += 1

                elapsed = (time.time() - start_timer)

                activity = str(index_calculated_paths) + \
                    " paths calculation from " + \
                    str(total_paths) + " total paths "

                self._report_time(elapsed, activity)

    def _report_time(self, time_spend, activity):
        print "{} took {:.2} seconds".format(activity, time_spend)

    def _store_network_paths(self, wb, gauge, ws_all=None):
        """Copy paths to general worksheet and to gauge specific worksheet."""

        # create worksheet and write field names
        ws = wb.create_sheet()
        ws.title = gauge + "_" + self.PATH_SHEET_SUFFIX
        ws.append(self.PATH_FIELDS)

        # take paths of the gauge
        paths = self.paths[gauge]

        # take nodes to iterate
        nodes = sorted(paths.keys())

        # iterate throught paths, copying them to worksheet
        for node_a in nodes:
            for node_b in nodes:

                # take distance
                distance = paths[node_a][node_b]["distance"]

                # create string por path
                list_path = paths[node_a][node_b]["path"]

                if list_path:
                    list_path = [str(node).zfill(3) for node in list_path]
                    string_path = "-".join(list_path)

                else:
                    string_path = None

                # create link name
                link_name = str(node_a) + "-" + str(node_b)

                # copy data to worksheet
                data = [link_name, node_a, node_b, distance, string_path]
                ws.append(data)

                # copy data to global worksheet
                if ws_all:
                    data.append(gauge)
                    ws_all.append(data)

    def _sort_by_distance(self, ws):
        """Sort paths in a worksheet by distance."""
        pass


def main(xl_input, gauge_names, xl_output):
    """Find shortest paths between all nodes of a network, by gauge.

    Args:
        xl_input: List of links of a network, by gauge.
        gauge_names: Names of the different gauges that are in the network.
        xl_output: List of shortest paths between all nodes, by gauge.
    """

    # load list of links
    wb = load_workbook(xl_input)

    # create a Network object
    network = Network(gauge_names)

    # create graphs from links, find shortest paths and store then in excel
    network.create_graphs_from_excel(wb)
    network.calculate_paths()
    network.store_paths_in_excel(wb, xl_output)

    # return network object, in case of the user wants to use it
    return network


def main_railway():
    """Find shortest paths for the railway network."""

    XL_INPUT = "data/railway_links_table.xlsx"
    GAUGE_NAMES = ["ancha", "media", "angosta"]
    XL_OUTPUT = "paths/railway_shortest_paths.xlsx"
    network = main(XL_INPUT, GAUGE_NAMES, XL_OUTPUT)

    return network


def main_roadway():
    """Find shortest paths for the roadway network."""

    XL_INPUT = "data/roadway_links_table.xlsx"
    GAUGE_NAMES = ["unica"]
    XL_OUTPUT = "paths/roadway_shortest_paths.xlsx"
    network = main(XL_INPUT, GAUGE_NAMES, XL_OUTPUT)

    return network


if __name__ == "__main__":

    # parse arguments if called with arguments
    if len(sys.argv) == 4:
        xl_input = sys.argv[1]
        gauge_names = sys.argv[2].split("-")
        xl_output = sys.argv[3]

        main(xl_input, gauge_names, xl_output)

    # call methods using default arguments if none are passed
    else:
        main_railway()
        main_roadway()
